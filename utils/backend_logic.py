# # utils/backend_logic.py
# import pandas as pd
# import openpyxl
# from openpyxl.utils.dataframe import dataframe_to_rows
# import datetime
# import io
# import traceback
# import sys
# import os
# import base64

# # This block ensures Python can find your 'main_models' package
# project_root = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
# if project_root not in sys.path:
#     sys.path.append(project_root)

# # This is the exact import block from your working revenue_logic.py script
# IMPORTS_OK = False
# IMPORT_ERROR_MESSAGE = ""
# try:
#     from main_models.imbalance_algorithm_SAP import run_battery_trading as run_battery_trading_SAP
#     from main_models.self_consumption_PV_PAP import run_battery_trading as run_battery_trading_PAP
#     from main_models.day_ahead_trading_PAP import run_battery_trading as run_battery_trading_day_ahead
#     from main_models.imbalance_everything_PAP import run_battery_trading as run_battery_trading_everything_PAP
#     IMPORTS_OK = True
# except ImportError as e:
#     IMPORT_ERROR_MESSAGE = f"Critical Error: Could not import an algorithm file from 'main_models'. Please check file and function names. Details: {e}"

# def run_master_simulation(params, input_df, progress_callback):
#     """
#     This is your original revenue_logic function, fully adapted for Dash.
#     """
#     if not IMPORTS_OK:
#         return {"summary": None, "warnings": [], "error": IMPORT_ERROR_MESSAGE}

#     now = datetime.datetime.now()
#     warnings = []
    
#     class Cfg: pass
#     config = Cfg()
#     for k, v in params.items():
#         setattr(config, k, v)
    
#     config.input_data = input_df

#     progress_callback("Starting model run...")

#     try:
#         strategy = params["STRATEGY_CHOICE"]
        
#         if strategy == "Simple Battery Trading (Imbalance)":
#             df, summary = run_battery_trading_SAP(config, progress_callback=progress_callback)
#         elif strategy == "Advanced Whole-System Trading (Imbalance)":
#             df, summary = run_battery_trading_everything_PAP(config, progress_callback=progress_callback)
#         elif strategy == "Optimize on Day-Ahead Market":
#             df, summary = run_battery_trading_day_ahead(config, progress_callback=progress_callback)
#         elif strategy == "Prioritize Self-Consumption":
#             df, summary = run_battery_trading_PAP(config, progress_callback=progress_callback)
#         else:
#             raise ValueError(f"Unknown strategy: {strategy}")

#         if df is None or not isinstance(df, pd.DataFrame):
#             raise ValueError("Model run failed to return a valid DataFrame.")

#         progress_callback("Model run complete. Generating Excel output...")
        
#         if strategy == "Simple Battery Trading (Imbalance)":
#             desired_columns = ['regulation_state', 'price_surplus', 'price_shortage', 'price_day_ahead', 'space_for_charging_kWh', 'space_for_discharging_kWh', 'energy_charged_kWh', 'energy_discharged_kWh', 'SoC_kWh', 'SoC_pct', 'grid_exchange_kWh', 'e_program_kWh', 'day_ahead_result', 'imbalance_result', 'energy_tax', 'supplier_costs', 'transport_costs', 'total_result_imbalance_SAP']
#         elif strategy == "Advanced Whole-System Trading (Imbalance)":
#             desired_columns = ['regulation_state', 'price_surplus', 'price_shortage', 'price_day_ahead', 'space_for_charging_kWh', 'space_for_discharging_kWh', 'energy_charged_kWh', 'energy_discharged_kWh', 'SoC_kWh', 'SoC_pct', 'grid_exchange_kWh', 'e_program_kWh', 'day_ahead_result', 'imbalance_result', 'energy_tax', 'supplier_costs', 'transport_costs', 'total_result_imbalance_PAP']
#         elif strategy == "Optimize on Day-Ahead Market":
#             desired_columns = ['production_PV', 'load', 'grid_exchange_kWh', 'price_day_ahead', 'space_for_charging_kWh', 'space_for_discharging_kWh', 'energy_charged_kWh', 'energy_discharged_kWh', 'SoC_kWh', 'SoC_pct', 'dummy1', 'dummy2', 'day_ahead_result', 'dummy3', 'energy_tax', 'supplier_costs', 'transport_costs', 'total_result_day_ahead_trading']
#         else:
#             desired_columns = ['production_PV', 'load', 'grid_exchange_kWh', 'price_day_ahead', 'space_for_charging_kWh', 'space_for_discharging_kWh', 'energy_charged_kWh', 'energy_discharged_kWh', 'SoC_kWh', 'SoC_pct', 'dummy1', 'dummy2', 'day_ahead_result', 'dummy3', 'energy_tax', 'supplier_costs', 'transport_costs', 'total_result_self_consumption']

#         df.index.name = 'Datetime'
#         existing_columns = [col for col in desired_columns if col in df.columns]
#         df_export = df[existing_columns]
        
#         wb = openpyxl.Workbook()
#         ws = wb.active
#         ws.title = "Import uit Python"
        
#         rows = dataframe_to_rows(df_export, index=True, header=True)
#         for r_idx, row in enumerate(rows, start=7):
#             for c_idx, value in enumerate(row, start=2):
#                 ws.cell(row=r_idx, column=c_idx, value=value)
        
#         ws.merge_cells('C6:M6')
#         datum_str = now.strftime('%d-%m-%Y %Hu%M')
#         optimization_method = summary.get('optimization_method', 'Pyomo optimalisatie')
#         summary_text = (f"Python run {datum_str}     {params['POWER_MW']} MW     {params['CAPACITY_MWH']} MWh     {round(summary.get('total_cycles', 0), 1)} cycli per jaar.     Algoritme: {params['STRATEGY_CHOICE']}     Optimalisatie: {optimization_method}")
#         ws['C6'] = summary_text
        
#         ws['W2'] = params['POWER_MW']; ws['W3'] = params['CAPACITY_MWH']; ws['W4'] = params['MIN_SOC']; ws['W5'] = params['MAX_SOC']; ws['W6'] = params['EFF_CH']; ws['W7'] = params['EFF_DIS']; ws['W8'] = params['SUPPLY_COSTS']; ws['W9'] = params['TRANSPORT_COSTS']

#         output_buffer = io.BytesIO()
#         wb.save(output_buffer)
#         output_buffer.seek(0)
        
#         if 'warning_message' in summary and summary['warning_message']:
#             warnings.append(summary['warning_message'])
#         if 'infeasible_days' in summary and len(summary.get('infeasible_days', [])) > 0:
#             warnings.append(f"Model was infeasible for {len(summary['infeasible_days'])} days.")

#         progress_callback("Output file generated successfully!")
        
#         file_bytes_b64 = base64.b64encode(output_buffer.getvalue()).decode('utf-8')

#         return {
#             "df": df.to_json(orient='split'),
#             "summary": summary,
#             "output_file_bytes_b64": file_bytes_b64,
#             "warnings": warnings,
#             "error": None
#         }

#     except Exception as e:
#         tb = traceback.format_exc()
#         print(f"ERROR in backend_logic: {tb}")
#         return {"summary": None, "warnings": [], "error": f"An error occurred during the model run: {str(e)}"}

# utils/backend_logic.py
import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import datetime
import io
import traceback
import sys
import os
import base64

# This block ensures Python can find your 'main_models' package
project_root = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
if project_root not in sys.path:
    sys.path.append(project_root)

# This is the exact import block from your working revenue_logic.py script
IMPORTS_OK = False
IMPORT_ERROR_MESSAGE = ""
try:
    from main_models.imbalance_algorithm_SAP import run_battery_trading as run_battery_trading_SAP
    from main_models.self_consumption_PV_PAP import run_battery_trading as run_battery_trading_PAP
    from main_models.day_ahead_trading_PAP import run_battery_trading as run_battery_trading_day_ahead
    from main_models.imbalance_everything_PAP import run_battery_trading as run_battery_trading_everything_PAP
    IMPORTS_OK = True
except ImportError as e:
    IMPORT_ERROR_MESSAGE = f"Critical Error: Could not import an algorithm file from 'main_models'. Please check file and function names. Details: {e}"

def run_master_simulation(params, input_df, progress_callback):
    """
    This is your original revenue_logic function, fully adapted for Dash.
    """
    if not IMPORTS_OK:
        return {"summary": None, "warnings": [], "error": IMPORT_ERROR_MESSAGE}

    now = datetime.datetime.now()
    warnings = []
    
    # Use a simple class to mimic the config object your models expect
    class Cfg: pass
    config = Cfg()
    for k, v in params.items():
        setattr(config, k, v)
    
    config.input_data = input_df

    progress_callback("Starting model run...")

    try:
        strategy = params["STRATEGY_CHOICE"]
        
        # This logic from your script correctly calls the different models
        if strategy == "Simple Battery Trading (Imbalance)":
            df, summary = run_battery_trading_SAP(config, progress_callback=progress_callback)
        elif strategy == "Advanced Whole-System Trading (Imbalance)":
            df, summary = run_battery_trading_everything_PAP(config, progress_callback=progress_callback)
        elif strategy == "Optimize on Day-Ahead Market":
            df, summary = run_battery_trading_day_ahead(config, progress_callback=progress_callback)
        elif strategy == "Prioritize Self-Consumption":
            df, summary = run_battery_trading_PAP(config, progress_callback=progress_callback)
        else:
            raise ValueError(f"Unknown strategy: {strategy}")

        if df is None or not isinstance(df, pd.DataFrame):
            raise ValueError("Model run failed to return a valid DataFrame.")

        progress_callback("Model run complete. Generating Excel output...")
        
        # --- This is your complete logic for creating the Excel file ---
        if strategy == "Simple Battery Trading (Imbalance)":
            desired_columns = [
                'regulation_state', 'price_surplus', 'price_shortage', 'price_day_ahead',
                'space_for_charging_kWh', 'space_for_discharging_kWh', 'energy_charged_kWh',
                'energy_discharged_kWh', 'SoC_kWh', 'SoC_pct', 'grid_exchange_kWh',
                'e_program_kWh', 'day_ahead_result', 'imbalance_result', 'energy_tax',
                'supplier_costs', 'transport_costs', 'total_result_imbalance_SAP'
            ]
        elif strategy == "Advanced Whole-System Trading (Imbalance)":
            desired_columns = [
                'regulation_state', 'price_surplus', 'price_shortage', 'price_day_ahead',
                'space_for_charging_kWh', 'space_for_discharging_kWh', 'energy_charged_kWh',
                'energy_discharged_kWh', 'SoC_kWh', 'SoC_pct', 'grid_exchange_kWh',
                'e_program_kWh', 'day_ahead_result', 'imbalance_result', 'energy_tax',
                'supplier_costs', 'transport_costs', 'total_result_imbalance_PAP'
            ]
        elif strategy == "Optimize on Day-Ahead Market":
            desired_columns = [
                'production_PV', 'load', 'grid_exchange_kWh', 'price_day_ahead',
                'space_for_charging_kWh', 'space_for_discharging_kWh', 'energy_charged_kWh',
                'energy_discharged_kWh', 'SoC_kWh', 'SoC_pct', 'dummy1', 'dummy2',
                'day_ahead_result', 'dummy3', 'energy_tax', 'supplier_costs',
                'transport_costs', 'total_result_day_ahead_trading'
            ]
        else:  # For "Prioritize Self-Consumption"
            desired_columns = [
                'production_PV', 'load', 'grid_exchange_kWh', 'price_day_ahead',
                'space_for_charging_kWh', 'space_for_discharging_kWh', 'energy_charged_kWh',
                'energy_discharged_kWh', 'SoC_kWh', 'SoC_pct', 'dummy1', 'dummy2',
                'day_ahead_result', 'dummy3', 'energy_tax', 'supplier_costs',
                'transport_costs', 'total_result_self_consumption'
            ]

        df.index.name = 'Datetime'
        existing_columns = [col for col in desired_columns if col in df.columns]
        df_export = df[existing_columns]
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Import uit Python"
        
        rows = dataframe_to_rows(df_export, index=True, header=True)
        for r_idx, row in enumerate(rows, start=7):
            for c_idx, value in enumerate(row, start=2):
                ws.cell(row=r_idx, column=c_idx, value=value)
        
        ws.merge_cells('C6:M6')
        datum_str = now.strftime('%d-%m-%Y %Hu%M')
        optimization_method = summary.get('optimization_method', 'Pyomo optimalisatie')
        summary_text = (
            f"Python run {datum_str}     {params['POWER_MW']} MW     {params['CAPACITY_MWH']} MWh     "
            f"{round(summary.get('total_cycles', 0), 1)} cycli per jaar.     "
            f"Algoritme: {params['STRATEGY_CHOICE']}     "
            f"Optimalisatie: {optimization_method}"
        )
        ws['C6'] = summary_text
        
        ws['W2'] = params['POWER_MW']
        ws['W3'] = params['CAPACITY_MWH']
        ws['W4'] = params['MIN_SOC']
        ws['W5'] = params['MAX_SOC']
        ws['W6'] = params['EFF_CH']
        ws['W7'] = params['EFF_DIS']
        ws['W8'] = params['SUPPLY_COSTS']
        ws['W9'] = params['TRANSPORT_COSTS']

        output_buffer = io.BytesIO()
        wb.save(output_buffer)
        output_buffer.seek(0)
        
        if 'warning_message' in summary and summary['warning_message']:
            warnings.append(summary['warning_message'])
        if 'infeasible_days' in summary and len(summary.get('infeasible_days', [])) > 0:
            warnings.append(f"Model was infeasible for {len(summary['infeasible_days'])} days.")

        progress_callback("Output file generated successfully!")
        
        # --- Final formatting for Dash ---
        file_bytes_b64 = base64.b64encode(output_buffer.getvalue()).decode('utf-8')

        return {
            "df": df.to_json(orient='split'),
            "summary": summary,
            "output_file_bytes_b64": file_bytes_b64,
            "warnings": warnings,
            "error": None
        }

    except Exception as e:
        tb = traceback.format_exc()
        print(f"ERROR in backend_logic: {tb}")
        return {"summary": None, "warnings": [], "error": f"An error occurred during the model run: {str(e)}"}